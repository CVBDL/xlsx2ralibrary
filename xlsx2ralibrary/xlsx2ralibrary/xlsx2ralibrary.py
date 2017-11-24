import argparse
import logging
import openpyxl
import os
import requests
import time
import urllib3


# Suppress https certificate warning.
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Setup logger.
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger('xlsx2ralibrary')

# https certificate file path
certificate_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                'assets',
                                'certificate.cer')

def parse_cli_args():
    """Process command line parameters.

    Returns:
        A dict contains 'user_name', 'password' and 'path'.
    Raises:
        Exception: Missing some required parameters.
    """
    parser = argparse.ArgumentParser(description='Import books to RaLibrary.')
    parser.add_argument('--user-name', help='RA-INT account name without ra-int perfix.')
    parser.add_argument('--password', help='RA-INT account password.')
    parser.add_argument('--path', help='Input file path.')
    args = parser.parse_args()
    # validate authentication info
    if not args.user_name or not args.password:
        raise Exception('[Abort] Missing accout name or password')
    # validate existence of input excel file
    if not args.path:
        raise Exception('[Abort] Missing input file path.')
    return args

def login(username, password):
    """Authenticate with INT domain account.

    Args:
        username (str): Domain user name.
        password (str): The password.
    Returns:
        str: An id token indicates the user's identity.
    Raises:
        Exception: Authentication failed.
    """
    endpoint = r'https://apcndaec3ycs12.ra-int.com/raauthentication/api/user'
    payload = { 'UserName': username, 'Password': password }
    req = requests.post(endpoint, data=payload, verify=certificate_path)
    if req.status_code == 200:
        return req.json()['IdToken']
    else:
        raise Exception('[Abort] Unauthorized.')

def get_authorization_header(id_token):
    """Generate authorization http request header.

    Args:
        id_token (str): The id token.
    Returns:
        A dict contains 'Authorization'.
    Raises:
        Exception: No id token provided.
    """
    if not id_token:
        raise Exception('Unauthorized.')
    return { 'Authorization': 'Bearer ' + id_token }

def query_book(isbn):
    """Query books details via books open API.

    Args:
        isbn (str): Book's ISBN.
    Returns:
        A dict of book.
    Raises:
        Exception: Query book failed.
    """
    if not isbn or not isinstance(isbn, str):
        raise Exception
    endpoint = r'https://apcndaec3ycs12.ra-int.com/ralibrary/api/book/isbn/'
    query_endpoint = '{0}{1}'.format(endpoint, isbn)
    req = requests.get(query_endpoint, verify=certificate_path)
    if req.status_code == 200:
        return req.json()
    else:
        raise

def save_book(book, headers):
    """Save book to RaLibrary.

    Args:
        book (dict): Book's data.
        headers (dict): HTTP header contains 'Authorization'.
    Raises:
        Exception: Saving book failed.
    """
    endpoint = r'https://apcndaec3ycs12.ra-int.com/ralibrary/api/books'
    req = requests.post(endpoint, headers=headers, data=book, verify=certificate_path)
    if req.status_code != 201:
        raise Exception(req.text)

def read_excel_rows(file_path):
    """Read excel rows and skip the first header row.

    Args:
        file_path (str): Input excel file path.
    Returns:
        Iterable rows.
    """
    wb = openpyxl.load_workbook(filename=file_path, read_only=True)
    ws = wb.active
    rows = iter(ws.rows)
    # skip header row
    next(rows)
    return rows

def process_row(row):
    """Parse a row in excel to book model.

    Args:
        row: A row in excel.
    Returns:
        A dict of book data.
    """
    isbn, code, title = map(lambda cell: str(cell.value), row)
    if not isbn or not code or not title:
        pass
    book = {}
    try:
        book = query_book(isbn)
    except:
        if len(isbn) == 10:
            isbn_10 = isbn
            isbn_13 = ''
        else:
            isbn_10 = ''
            isbn_13 = isbn
        book['ISBN10'] = isbn_10
        book['ISBN13'] = isbn_13
        book['Title'] = title
    # 'Code' must be manually assigned.
    book['Code'] = code
    return book

def main():
    """Main function to begin the processing."""
    # Parse command line arguments.
    try:
        logger.info('Parsing command line parameters...')
        args = parse_cli_args()
        logger.info('Parsing command line parameters...Successfully')
    except Exception as e:
        logger.info(e);
        return 1

    # Authentication.
    try:
        logger.info('Verifying your credentials...')
        id_token = login(args.user_name, args.password)
        logger.info('Verifying your credentials...Successfully')
    except Exception as e:
        logger.info(e);
        return 1

    # Read input books data.
    try:
        logger.info('Reading Excel file...')
        rows = read_excel_rows(args.path)
        logger.info('Reading Excel file...Successfully')
    except:
        logger.info('[Abort] Cannot read the input Excel file.')
        return 1

    # Create books to RaLibrary.
    try:
        # Create a new worksheet inside input Excel file for logging purpose.
        wb = openpyxl.load_workbook(filename=args.path)
        ws_name = 'FailedAddedBooks_{0}'.format(str(time.time())[0:10])
        ws_log = wb.create_sheet(title=ws_name)
        # Generate authorization HTTP header.
        authorization_header = get_authorization_header(id_token)
        # Start to create books to RaLibrary.
        logger.info('Creating books to RaLibrary...')
        failureCount = 0
        for row in rows:
            book = process_row(row)
            try:
                save_book(book, authorization_header)
            except Exception as e:
                failureCount += 1
                log_row = list(map(lambda cell: str(cell.value), row))
                log_row.append(str(e))
                ws_log.append(log_row)
        wb.save(filename=args.path)
        logger.info('Creating books to RaLibrary...Done')
        logger.info('Total failures: {0}'.format(failureCount))
        logger.info('View the logs:')
        logger.info('    File: {0}'.format(args.path))
        logger.info('    Worksheet: {0}'.format(ws_name))
    except Exception as e:
        logger.info(e)
        return 1

    return 0
